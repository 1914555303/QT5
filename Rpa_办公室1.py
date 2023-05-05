import json
from configparser import ConfigParser
import pyautogui as gui
import time
import pyperclip
import sys
import os
import requests
import datetime
import re
from openpyxl import load_workbook
import logging

CONFIGFILE = '.\\config\\config1.ini'
config = ConfigParser()
config.read(CONFIGFILE)
Click01 = config.get('click01', 'po').replace('_', ' ')
Click02 = config.get('click02', 'po').replace('_', ' ')
Click03 = config.get('click03', 'po').replace('_', ' ')
Click04 = config.get('click04', 'po').replace('_', ' ')
Click05 = config.get('click05', 'po').replace('_', ' ')
Click06 = config.get('click06', 'po').replace('_', ' ')
Click07 = config.get('click07', 'po').replace('_', ' ')
Click08 = config.get('click08', 'po').replace('_', ' ')
Click09 = config.get('click09', 'po').replace('_', ' ')
Click10 = config.get('click10', 'po').replace('_', ' ')
Click11 = config.get('click11', 'po').replace('_', ' ')

log_path = os.getcwd() + '\\log'
if not os.path.exists(log_path):
    os.makedirs(log_path)
logging.basicConfig(level=logging.ERROR,
                    filename='./log/log.txt',
                    filemode='a+',
                    format='%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s')


def Clicks(clicks):
    time.sleep(0.2)
    z = clicks.split()
    xx = int(z[0])
    yy = int(z[1])
    gui.click(xx, yy)
    time.sleep(0.2)

def Clicks_R(clicks):
    time.sleep(0.2)
    z = clicks.split()
    xx = int(z[0])
    yy = int(z[1])
    gui.click(xx, yy, button='right')
    time.sleep(0.2)

now_times = datetime.datetime.now()
now_data = now_times.strftime('%Y.%m.%d')
now_time = now_times.strftime('%H.%M.%S')
# db = pymssql.connect(host='10.10.100.125',  user='hrq', password='Hc2673888', database='HCDB01')
# curses = db.cursor()

# 检测屏幕分辨率是否符合要求
def check_rp():
    time.sleep(2)
    w, h = 1920, 1080
    nwe_w, nwe_h = gui.size()
    print(nwe_w, nwe_h)
    if w != nwe_w or h != nwe_h:
        print('当前屏幕分辨率为：%s,%s,该电脑屏幕分辨率不符合要求，请在分辨率为1920*1080的电脑上运行!' % (nwe_w, nwe_h))
        os.system("pause")
        sys.exit()
    else:
        print('屏幕分辨率符合要求！')


#
def send_post(touser, message):
    send_url = 'http://10.10.100.126:5001/testsendmsg/'+touser
    respone = requests.post(send_url, data={'msg': message})

def copy(num):
    time.sleep(0.7)
    pyperclip.copy(num)  # 复制num的值
    gui.hotkey('ctrl', 'v')
    time.sleep(0.7)

def Tab(n):
    for k in range(n):
        time.sleep(0.2)
        gui.press('tab')

def error_ex(hdh):
    time.sleep(1)
    sql = "UPDATE dbo.cgnxjhd SET flag = '3' where nxjhdh= %s " % hdh
    updates = requests.post('http://10.10.20.30:8443/exejsonsql', data={"sql": sql}).text
    msg = '内向交货单：{}\nSAGE制单未成功'.format(hdh)
    send_post('20743', msg)
    os.system("pause")
    sys.exit()

# 登录
def sage_login():
    time.sleep(2)
    Clicks(Click01)
    time.sleep(5)
    gui.press('enter')
    time.sleep(7)

def sage_out():
    time.sleep(2)
    gui.hotkey('alt', 'f4')
    time.sleep(0.5)
    gui.press('enter')

# 选择订单类型
def select_type(n):  # 45是标准，46是寄售
    time.sleep(2)
    Clicks(Click02)

    time.sleep(1)
    if n == '45':
        gui.press('down')
    else:
        gui.press('down')
        time.sleep(0.5)
        gui.press('down')
    time.sleep(1)
    gui.press('enter')


def creat(gong_num, beizhus):
    time.sleep(5)
    gui.hotkey('ctrl', 'n')
    time.sleep(1)

    gui.press('tab')
    time.sleep(0.5)
    gui.press('tab')
    copy(gong_num)  # 给粘贴板一个供应商号
    gui.press('tab')
    time.sleep(1)
    if gui.getActiveWindowTitle() == '信息消息':
        gui.press('enter')
    time.sleep(0.5)
    Tab(3)
    copy(beizhus)


# 输入物料
def input_m(list_mm, list_s):
    time.sleep(2)
    Clicks(Click03)                         ###############行图标
    Clicks(Click03)                         ###############行图标
    gui.press('tab')

    for i in range(0, len(list_mm)):
        copy(list_mm[i])  # 物料号
        gui.press('enter')
        time.sleep(1)

        ########
        fw = gui.getActiveWindowTitle()
        if fw == "待考虑需求":
            gui.press('enter')
        if fw == '字段错误 "采购单位"':
            gui.press('enter')
            list_cgdw = ['ZHI', 'TAO', 'PCS']
            for s in range(0, 3):
                time.sleep(1)
                copy(list_cgdw[s])
                gui.press('enter')
                if gui.getActiveWindowTitle() == '字段错误 "已订购"':
                    break
                else:
                    time.sleep(1)
                    gui.press('enter')

        time.sleep(1)
        gui.press('enter')

        copy(list_s[i])
        gui.press('enter')
        time.sleep(1)
        if gui.getActiveWindowTitle() == "警告":
            gui.press('enter')

        time.sleep(1)
        gui.press('down')
        for j in range(6):
            time.sleep(0.1)
            gui.press('left')
        time.sleep(0.5)
        gui.press('1')
        time.sleep(0.5)
        gui.press('backspace')
    time.sleep(0.5)
    gui.press('esc')


def print_sage(nn):
    time.sleep(1)
                               #####################新建按钮
    Clicks(Click04)
    Clicks(Click04)
    if nn > 10:
        time.sleep(10)
    else:
        time.sleep(3)
    fw = gui.getActiveWindowTitle()
    ret = re.findall('日志..读取', fw)
    if ret != []:
        time.sleep(0.5)
        time.sleep(0.5)
                             ############# 发生错误的结束按钮
        Clicks(Click05)

    time.sleep(1)
    # 获取订单号到粘贴板
    gui.hotkey('ctrl', 'c')
    time.sleep(0.5)
    a = pyperclip.paste()  # 粘贴单据号

    if a[0:2] != 'PO':
        res = f'result\{a}+1.png'  # 截图并保存
        gui.screenshot(res)
        os.system("pause")

    time.sleep(2)
    res = f'result\{a}+2.png'  # 截图并保存
    gui.screenshot(res)

    return a

# 打印并结束
def end(hdh):
    time.sleep(1.5)
                                        ################# 打印图标
    Clicks(Click06)
    time.sleep(1)
    gui.press('enter')

    time.sleep(1.5)
    gui.press('enter')
    time.sleep(3)

    Clicks(Click07)
    time.sleep(1)
    gui.hotkey('win', 'up')

    time.sleep(1.5)
    Clicks(Click08)
    time.sleep(1.5)
    gui.press('enter')
    time.sleep(1.5)

    # Clicks(Click09)
    # time.sleep(1)
    gui.hotkey('alt', 'f4')
    time.sleep(2)
    Clicks(Click10)
    time.sleep(2)
    ####
    time.sleep(1)
    fw1 = gui.getActiveWindowTitle()
    if fw1[0:9] != 'HZGEARBOX':
        print("打印过程出现错误，请到sage查看打印情况，若进行手动打印此条入库单，完成后将flag改为4")
        error_ex(hdh)

# 检查
def check_h():
    time.sleep(1.5)
    Clicks(Click11)
    time.sleep(1)
    Clicks_R(Click11)
    time.sleep(1)
    for i in range(12):
        gui.press('down')
        time.sleep(0.1)
    time.sleep(0.5)
    gui.press('enter')
    a = pyperclip.paste()
    with open("save_sage.txt", 'w', encoding='utf-8') as f:
        f.write(a)
    with open("save_sage.txt", 'r+', encoding='utf-8') as file:
        lines = file.readlines()
    return len(lines)

def get_Userid():
    wb = load_workbook(".\\config\\办公室1.xlsx")
    ws = wb['Sheet1']
    list_id = []
    rows = ws.max_row
    for i in range(0, rows):
        if str(ws[f'B{i + 2}'].value) == 'None' or str(ws[f'B{i + 2}'].value) == '':
            continue
        list_id.append(str(ws[f'B{i + 2}'].value))
    wb.close()
    tj = ""
    for id in list_id:
        tj = tj + f"userid='{id}'"
        if id != list_id[-1]:
            tj = tj + ' or '
    return tj

# def test813():
#     time.sleep(3)
#     ###############################         主表，取nxjhdh
#     sql = f"select nxjhdh, userid, sl from dbo.cgnxjhd where flag='2' and ({tjs})"
#     curses.execute(sql)
#     dh = curses.fetchone()
#     if dh is None:
#         nxjhdh = []
#         userid = []
#         ts = []
#     else:
#         nxjhdh = [dh[0]]
#         userid = dh[1]
#         ts = int(dh[2])
#
#     print("将要录入的交货单号有：%s" % nxjhdh)
#     ######
#     for dh in nxjhdh:
#         sql = "select cgddh,sagegysbm from dbo.cgnxjhdmx where nxjhdh= %s " % dh
#         curses.execute(sql)
#         gongs = set(curses.fetchall())
#         sql2 = "UPDATE dbo.cgnxjhd SET flag = '3' where nxjhdh= %s " % dh  # 正在录入的单号
#         curses.execute(sql2)
#         db.commit()
#         print("正在进行录入的交货单号是：%s" % dh)
#
#         for gong in gongs:
#             cgddh = gong[0]  # 采购订单号
#             gon = gong[1]  # 供应商
#             print("目前供应商号是：%s" % gon)
#             beizhu = "nxjhdh:%s,cgddh:%s" % (dh, cgddh)  # 备注
#             cgd = cgddh[0:2]  # 判断采购订单号前两位的值45/46
#
#             select_type(cgd)  ##### 选择订单类型
#
#             creat(gon, beizhu)  ##### 创建
#
#             sql1 = "select wlm,sl,sagewlm from dbo.cgnxjhdmx where sagegysbm= %s and nxjhdh = %s " % (gon, dh)
#             curses.execute(sql1)
#             data_alls = curses.fetchall()
#             list_wl = []
#             list_sl = []
#             for data_all in data_alls:
#                 if data_all[2] == '':
#                     list_wl.append(data_all[0])
#                 else:
#                     list_wl.append(data_all[2])
#                 # list_sl.append((str(data_all[1])[:-4]).replace(',', ''))
#                 # list_sl.append(data_all[1])
#                 aa = data_all[1].split('.')
#                 try:
#                     if int(aa[1]) != 0:
#                         list_sl.append(data_all[1].replace(',', ''))
#                     else:
#                         list_sl.append(aa[0].replace(',', ''))
#                 except:
#                     list_sl.append(data_all[1].replace(',', ''))
#
#
# ####################################################
#             time.sleep(0.5)
#             try:
#                 input_m(list_wl, list_sl)
#             except:
#                 print("物料录入错误，请到sage检查入库单录入情况")
#                 error_ex(dh)
#
#             time.sleep(0.5)
#             djh = print_sage(len(list_wl))  ###### 确认建立，并返回单据号
#             print("单据号为：%s" % djh)
#             sql = "UPDATE dbo.cgnxjhdmx SET sagedjh = %s WHERE sagegysbm= %s and nxjhdh = %s "  # 单据号回写数据库
#             val = (djh, gon, dh)
#             curses.execute(sql, val)
#             db.commit()
#
#             time.sleep(1)
#             lrts = check_h()
#
#             time.sleep(1)
#             end(dh)          ###### 创建成功，结束或返回进行下一个创建
#
#             sql3 = "UPDATE dbo.cgnxjhd SET flag = '4' where nxjhdh= %s " % dh
#             curses.execute(sql3)
#             db.commit()
#             msg = f'内向交货单：{dh}\nSAGE已制单成功：{djh}\n理论录入行数：{ts}\nSAGE实际录入行数：{lrts}'
#             send_post(f'20743|{userid}', msg)

def test813():
    # time.sleep(1)
    ###############################         主表，取nxjhdh
    sql = f"select nxjhdh, userid, sl from dbo.cgnxjhd where flag=2 and ({tjs})"
    # sql = f"select nxjhdh, userid, sl from dbo.cgnxjhd where nxjhdh='180025365'"
    datas = requests.post('http://10.10.20.30:8443/getjsonone', data={"sql": sql}).text
    datas_one = json.loads(datas)
    if datas_one is None:
        return

    else:
        nxjhdh = datas_one['nxjhdh']
        userid = datas_one['userid']
        ts = int(datas_one['sl'])
        dh = nxjhdh
    print("将要录入的交货单号有：%s" % nxjhdh)
    ######

    sql = f"select wlm,sl,sagewlm, cgddh,sagegysbm from dbo.cgnxjhdmx where nxjhdh={dh}"
    datas = requests.post('http://10.10.20.30:8443/getjsonall', data={"sql": sql}).text
    datas_all = json.loads(datas)

    sql2 = f"UPDATE dbo.cgnxjhd SET flag = '3' where nxjhdh= {dh}"   # 正在录入的单号
    updates = requests.post('http://10.10.20.30:8443/exejsonsql', data={"sql": sql2}).text
    print("正在进行录入的交货单号是：%s" % dh)


    cgddh = datas_all[0]['cgddh']  # 采购订单号
    gon = datas_all[0]['sagegysbm']  # 供应商

    list_wl = []
    list_sl = []
    for data_all in datas_all:
        if data_all["sagewlm"] == '':
            list_wl.append(data_all['wlm'])
        else:
            list_wl.append(data_all["sagewlm"])

        aa = data_all['sl'].split('.')
        try:
            if int(aa[1]) != 0:
                list_sl.append(data_all["sl"].replace(',', ''))
            else:
                list_sl.append(aa[0].replace(',', ''))
        except:
            list_sl.append(data_all["sl"].replace(',', ''))

    print("目前供应商号是：%s" % gon)

    beizhu = "nxjhdh:%s,cgddh:%s" % (dh, cgddh)  # 备注

    cgd = cgddh[0:2]  # 判断采购订单号前两位的值45/46


    print(f"物料码：{list_wl}")
    print(f"数量：{list_sl}")
    # time.sleep(10000)

####################################################
    select_type(cgd)  ##### 选择订单类型

    creat(gon, beizhu)  ##### 创建


    time.sleep(0.5)
    try:
        input_m(list_wl, list_sl)
    except:
        print("物料录入错误，请到sage检查入库单录入情况")
        error_ex(dh)

    time.sleep(0.5)
    djh = print_sage(len(list_wl))  ###### 确认建立，并返回单据号
    print("单据号为：%s" % djh)
    a = str(djh)
    if a[0: 2] != 'PO':
        time.sleep(2)
        msg = f'内向交货单：{dh}\nSAGE采购制单失败!!!'
        send_post(f'20743|{userid}', msg)
        os.system('pause')


    sql = f"UPDATE dbo.cgnxjhdmx SET sagedjh = '{djh}' WHERE nxjhdh = {dh} "  # 单据号回写数据库
    updates = requests.post('http://10.10.20.30:8443/exejsonsql', data={"sql": sql}).text

    time.sleep(1)
    lrts = check_h()

    time.sleep(1)
    end(dh)          ###### 创建成功，结束或返回进行下一个创建

    sql3 = "UPDATE dbo.cgnxjhd SET flag = '4' where nxjhdh= %s " % dh
    updates = requests.post('http://10.10.20.30:8443/exejsonsql', data={"sql": sql3}).text
    msg = f'内向交货单：{dh}\nSAGE已制单成功：{djh}\n理论录入行数：{ts}\nSAGE实际录入行数：{lrts}'
    send_post(f'20743|{userid}', msg)

def check():
    sql = f"select nxjhdh from dbo.cgnxjhd where flag='2' and ({tjs})"
    datas = requests.post('http://10.10.20.30:8443/getjsonone', data={"sql": sql}).text
    datas_one = json.loads(datas)
    if datas_one is None:
        nxjhdh = []
        return nxjhdh


if __name__ == '__main__':
    print("**********SAGE单据号录入RPA************")
    try:
        tjs = get_Userid()
        sage_login()
        x = 1
        while 2:
            time.sleep(2)
            if check() == []:
                if x == 1:
                    sage_out()
                    x = 0
            else:
                if x == 0:
                    sage_login()
                    x = 1
                    test813()
                else:
                    test813()
    except Exception as e:
        logging.exception(e)
        os.system("pause")

# sql = f"select userid, sl from dbo.cgnxjhdmx where nxjhdh='180022166'"
# curses.execute(sql)
# dh = curses.fetchall()
# list_sl = []
#
# for d in dh:
#     # print(d[1].replace('.', ''))
#     # print(int(float(d[1].replace(',', ''))))
#     # print(d[1].split('.'))
#     a = d[1].split('.')
#     # print(int(a[1]))
#     try:
#         if int(a[1]) != 0:
#             list_sl.append(d[1].replace(',', ''))
#         else:
#             list_sl.append(a[0].replace(',', ''))
#     except:
#         list_sl.append(d[1].replace(',', ''))
#
#     print(list_sl)
#
# print(dh)

# tjs = get_Userid()
# test813()

# msg = f'内向交货单：180025577\nSAGE已制单成功：POHHCG23050012\n理论录入行数：1\nSAGE实际录入行数：1'
# send_post(f'20743|17904', msg)
# djh = 'POHHCG23050020'
# dh = '180025593'
#
# sql = f"UPDATE dbo.cgnxjhdmx SET sagedjh = '{djh}' WHERE nxjhdh = {dh} "  # 单据号回写数据库
# updates = requests.post('http://10.10.20.30:8443/exejsonsql', data={"sql": sql}).text