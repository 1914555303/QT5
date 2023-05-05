import sys
from PyQt5 import QtWidgets,QtGui,QtCore,Qt

import os
import time
import pyautogui as gui
import pyperclip
import pymssql
import requests
import win32clipboard
from openpyxl import load_workbook

############################
db = pymssql.connect(host='10.10.100.125', user='hrq', password='Hc2673888', database='HCDB01', autocommit=True)
curses = db.cursor()

tuo = '20743'
def send_post(touser, message):
    send_url = 'http://10.10.100.126:5001/testsendmsg/'+touser
    respone = requests.post(send_url, data={'msg': message})

def copy(num):
    time.sleep(0.4)
    pyperclip.copy(num)  # 复制num的值
    time.sleep(0.1)
    gui.keyDown('ctrl')
    time.sleep(0.1)
    gui.keyDown('v')
    time.sleep(0.1)
    gui.keyUp('v')
    time.sleep(0.1)
    gui.keyUp('ctrl')
    time.sleep(0.4)

def sql_update(num, dh):
    sql = f"UPDATE dbo.tltask SET flag = {num} WHERE rkdh = '{dh}' "
    curses.execute(sql)
    db.commit()

def clickHui():
    time.sleep(1)
    for a in range(0, 3):
        time.sleep(0.5)
        gui.click(9, 118 + 17 * a)  # 灰圈
        time.sleep(1)
        fw = gui.getActiveWindowTitle()
        if fw == "X3应用错误":
            time.sleep(1)
            if gui.locateOnScreen("img/02.png") != None:
                gui.press('enter')
                break

def return_result(dh):
    time.sleep(1)
    # now_times = datetime.datetime.now()
    now_time = now_times.strftime('%H%M%S')
    File_Path = os.getcwd() + '\\错误截图\\{}\\'.format(now_data)
    if not os.path.exists(File_Path):
        os.makedirs(File_Path)
    res = '错误截图\\{}\\{}+{}.png'.format(now_data, dh, now_time)
    gui.screenshot(res)

def input_wl(wl):
    for x in range(5):
        time.sleep(0.2)
        copy(wl)
        win32clipboard.OpenClipboard()
        win32clipboard.EmptyClipboard()
        win32clipboard.CloseClipboard()
        time.sleep(0.5)
        gui.hotkey('ctrl', 'a')
        time.sleep(0.5)
        gui.hotkey('ctrl', 'c')
        qq = pyperclip.paste()
        if wl != qq:
            time.sleep(5)
            gui.hotkey('ctrl', 'a')
            time.sleep(0.5)
            gui.press('backspace')
        else:
            break

def select_KW(nnn):
    time.sleep(1.5)
    for i in range(3):
        time.sleep(2)
        cuo = gui.locateOnScreen("img/04.png")
        if cuo != None:
            gui.moveTo(cuo)
            break
    if cuo == None:
        msg = f"未识别到04.png\n{key}"
        send_post('20743', msg)
        return_result(f"未识别到04+{key}")
        gui.press('esc')
        sage_out()
        return 2

    gui.click(cuo, button='right')
    time.sleep(0.8)
    gui.press('down')
    time.sleep(0.5)
    gui.press('enter')
    time.sleep(0.5)
    gui.typewrite(nnn)
    time.sleep(0.5)
    gui.press('enter')

########################## sage
def sage_login():
    time.sleep(2)
    gui.click(90, 878)
    time.sleep(8)
    for k in range(8):
        time.sleep(0.2)
        gui.press('tab')
    copy('ZHONGGL1')
    gui.press('enter')
    time.sleep(12)

def sage_out():
    time.sleep(2)
    gui.hotkey('alt', 'f4')
    time.sleep(1)
    gui.press('enter')

def sage_de():
    time.sleep(1)
    gui.press('esc')
    time.sleep(0.5)
    gui.press('up')
    time.sleep(0.5)
    gui.press('numlock')
    time.sleep(0.5)
    gui.hotkey('shift', 'Delete')
    time.sleep(0.5)
    gui.press('numlock')
    time.sleep(0.5)
    gui.hotkey('shift', 'Delete')
    time.sleep(1)
    gui.doubleClick(770, 638)
############################

###  列表转字符
def list_to_str(lists):
    s = ""
    for i in lists:
        s += i.replace("'", '') + ','
    s = s[:-1]
    return s

###  保存sap过账数据（物料、数量、凭证号）
def sap_baocun(list_key, list_ylh, list_ms, list_wl, list_sl, dh, typ, scddh):
    dhs = f"{scddh}+{dh}"

    if typ == 3:
        ty = 'sage库存不足'
    if typ == 4:
        ty = 'sage库存不足副本'
    if typ == 5:
        ty = 'sage投料成功'

    File_Path = os.getcwd() + f'\\{ty}\\{now_data}\\{dhs}\\'
    if not os.path.exists(File_Path):
        os.makedirs(File_Path)

    with open(f"{ty}\\{now_data}\\{dhs}\\{dhs}+{key}.text", 'w', encoding='utf-8') as f:
        f.write(list_to_str(list_key))
        f.write('\n')
        f.write(list_to_str(list_ylh))
        f.write('\n')
        f.write(list_to_str(list_ms))
        f.write('\n')
        f.write(list_to_str(list_wl))
        f.write('\n')
        f.write(list_to_str(list_sl))

def check_fen(dhs):
    File_Path = os.getcwd() + f"\\分物料\\{now_data}\\{dhs}\\{dhs}+{key}.text"
    if not os.path.exists(File_Path):
        print(f"无{dhs}+{key}.text")
        time.sleep(10)
        return 1

    with open(f"分物料\\{now_data}\\{dhs}\\{dhs}+{key}.text", 'r+', encoding='utf-8') as f:
        a = f.readlines()
        if a[-1] != 'N':
            print(f"{dhs}+{key}.text已投料")
            time.sleep(10)
            return 1


def read_fen_wl(dhs):
    with open(f"分物料\\{now_data}\\{dhs}\\{dhs}+{key}.text", 'r+', encoding='utf-8') as f:
        a = f.readlines()
    list1_key = a[0].split(',')
    list1_key[-1] = list1_key[-1].strip()
    list1_ylh = a[1].split(',')
    list1_ylh[-1] = list1_ylh[-1].strip()
    list1_ms = a[2].split(',')
    list1_ms[-1] = list1_ms[-1].strip()

    list1_wls = a[3].split(',')
    list1_wls[-1] = list1_wls[-1].strip()
    list1_wl = []
    for line in list1_wls:
        list1_wl.append(line.lstrip("\ufeff"))
    list1_sl = a[4].split(',')
    list1_sl[-1] = list1_sl[-1].strip()

    return list1_key, list1_ylh, list1_ms, list1_wl, list1_sl

def ctrl_c(CTRL, C):
    gui.keyDown(CTRL)
    gui.keyDown(C)
    gui.keyUp(C)
    gui.keyUp(CTRL)

#### 总成投料
def touliao(list_key, list_ylh, list_ms, list_wl, list_sl, dh, scddh, typ):
    nn = '323'
    dhs = f"{scddh}+{dh}"
    time.sleep(2)
    gui.click(760, 168)  ## 其他出货
    time.sleep(2)
    for k1 in range(0, 4):
        gui.press('down')
    time.sleep(1)
    gui.press('enter')

    time.sleep(3)
    if gui.getActiveWindowTitle() == '信息消息':
        gui.press('enter')
    time.sleep(3)
    gui.hotkey('ctrl', 'n')
    copy(f'自动化投料过账{dh}+{key}')
    time.sleep(0.2)
    gui.press('tab')
    if typ == '1':
        copy('LHQ')
        nn = '382'
    gui.press('tab')
    times = now_times.strftime("%Y/%m/%d")[2:10]
    copy(times)
    for k3 in range(3):
        time.sleep(0.2)
        gui.press('tab')

    sage_wl_defect = []          ## sage库存不足物料
    sage_sl_defect = []
    sage_ylh_defect = []
    sage_ms_defect, sage_key_defect = [], []

    sage_wl_cz, sage_sl_cz, sage_ylh_cz = [], [], []      ## sage已投物料
    sage_ms_cz, sage_key_cz = [], []

    for i in range(0, len(list_wl)):

        input_wl(list_wl[i])
        gui.press('tab')

        time.sleep(1)
        fww = gui.getActiveWindowTitle()
        time.sleep(0.5)
        dongjie = gui.locateOnScreen("img/dongjie.png")
        if fww == '字段错误 "产品"' and dongjie != None:
            time.sleep(0.5)
            return_result('物料冻结')
            msg = f"入库单号：{dh}\n{list_wl[i]}物料冻结\n{key}"
            send_post('20743', msg)
            gui.press('enter')
            time.sleep(0.5)
            gui.press('backspace')
            sage_key_defect.append(list_key[i])
            sage_wl_defect.append(list_wl[i])
            sage_sl_defect.append(list_sl[i])
            sage_ylh_defect.append(list_ylh[i])
            sage_ms_defect.append(list_ms[i])

            if i == len(list_wl)-1:
                gui.press('esc')
                time.sleep(1)
                gui.click(1230, 828)  ## 新建
                break
            else:
                continue
        for k3 in range(2):
            gui.press('tab')

        time.sleep(0.3)
        # copy(list_sl[i])
        input_wl(list_sl[i])
        for k4 in range(0, 7):
            time.sleep(0.1)
            gui.press('tab')
        ##################
        if select_KW(nn) == 2:
            return 2

        clickHui()
        ##################
        time.sleep(1)
        gui.click(1170, 825)  ## 保存

        time.sleep(1.5)
        fw = gui.getActiveWindowTitle()
        if fw == "问题":
            time.sleep(1)
            gui.press('enter')
            sage_key_defect.append(list_key[i])
            sage_wl_defect.append(list_wl[i])
            sage_sl_defect.append(list_sl[i])
            sage_ylh_defect.append(list_ylh[i])
            sage_ms_defect.append(list_ms[i])

            if i != len(list_wl)-1:
                sage_de()

            else:
                time.sleep(1)
                gui.press('esc')
                time.sleep(1)
                gui.press('up')
                time.sleep(1)
                gui.press('numlock')
                time.sleep(0.5)
                gui.hotkey('shift', 'Delete')
                time.sleep(0.5)
                gui.press('numlock')
                time.sleep(0.5)
                gui.hotkey('shift', 'Delete')

                time.sleep(1)
                gui.click(1230, 828)  ## 新建

        else:
            sage_key_cz.append(list_key[i])
            sage_wl_cz.append(list_wl[i])
            sage_sl_cz.append(list_sl[i])
            sage_ylh_cz.append(list_ylh[i])
            sage_ms_cz.append(list_ms[i])

            if i == len(list_wl)-1:
                gui.press('esc')
                time.sleep(1)
                gui.click(1230, 828)  ## 新建


    sap_baocun(sage_key_cz, sage_ylh_cz, sage_ms_cz, sage_wl_cz, sage_sl_cz, dh, 5, scddh)
    if sage_wl_defect != []:
        sap_baocun(sage_key_defect, sage_ylh_defect, sage_ms_defect, sage_wl_defect, sage_sl_defect, dh, 3, scddh)

    time.sleep(6)
    gui.click(1230, 828)  ## 新建
    time.sleep(4)
    gui.click(1230, 828)  ## 新建
    time.sleep(6)
    gui.click(1230, 828)  ## 新建
    time.sleep(10)
    gui.click(892, 98)
    time.sleep(1)
    gui.mouseDown()
    gui.move(-150, 0, duration=0.7)
    gui.mouseUp()
    time.sleep(1)
    ctrl_c('ctrl', 'c')
    time.sleep(1)
    a = pyperclip.paste()  ## 拿到物料号
    # gui.hotkey('ctrl', 'c')
    # a = pyperclip.paste()
    #################

    #################
    if a[0: 6] == 'MITHCG':
        time.sleep(1)
        gui.click(1400, 828)    ## 结束

    else:
        msg = f"入库单号：{dh}\n生产订单号：{scddh}\nsage自动化投料过账失败\n{key}"
        send_post(tuo, msg)
        return_result(f"投料过账失败+{key}")
        with open(f"分物料\\{now_data}\\{dhs}\\{dhs}+{key}.text", 'a+', encoding='utf-8') as f:
            f.write('\n')
            f.write('Y')
        gui.press('esc')
        gui.press('esc')
        sage_out()
        return 2
        # os.system("pause")

    if sage_wl_defect != []:
        sap_baocun(sage_key_defect, sage_ylh_defect, sage_ms_defect, sage_wl_defect, sage_sl_defect, dh, 4, scddh)

def get_data():
    sql = "select rq from dbo.tltask where flag=2 "
    curses.execute(sql)
    rq = curses.fetchone()
    if rq is None:
        return 1
    else:
        data = rq[0]
        return data

def get_LHQ():
    wb = load_workbook(".\\config\\离合器总成清单.XLSX")
    ws = wb['Sheet1']
    list_lhq_wu = []
    for i in ws['A']:
        list_lhq_wu.append(i.value)
    list_lhq_wu.remove(list_lhq_wu[0])
    wb.close()

    return list_lhq_wu

def zzz():
    sql = "select rkdh from dbo.tltask where flag=2 "
    curses.execute(sql)
    dh = curses.fetchone()
    if dh is None:
        rkdh = []
    else:
        rkdh = [dh[0]]
    print(f"入库单号有：{rkdh}")

    for dh in rkdh:
        sql = "select wlm,scddh,gcbm,sl from dbo.tltask where rkdh = '{}' ".format(dh)
        curses.execute(sql)
        data_alls = curses.fetchall()
        print(data_alls)
        wlm, scddh, gc, sl = data_alls[0][0], data_alls[0][1], data_alls[0][2], data_alls[0][3]
        print(wlm, scddh, gc, sl)

        dhs = f"{scddh}+{dh}"
        tps = '2'
        for LHQ in list_LHQ:
            if LHQ == wlm:
                tps = '1'
                break
        # dhs = '110041293+R000024791'
###################################

        if check_fen(dhs) == 1:
            break

        else:
            # sage_login()
            list1_key, list1_ylh, list1_ms, list1_wl, list1_sl = read_fen_wl(dhs)

            if touliao(list1_key, list1_ylh, list1_ms, list1_wl, list1_sl, dh, scddh, tps) == 2:
                time.sleep(3)
                sage_login()
                break

            with open(f"分物料\\{now_data}\\{dhs}\\{dhs}+{key}.text", 'a+', encoding='utf-8') as f:
                f.write('\n')
                f.write('Y')

            # sage_out()







class GUI(QtWidgets.QWidget):
    def __init__(self):
        #初始化————init__
        super().__init__()
        self.initGUI()
        self.value = '1'
    def initGUI(self):
        #设置窗口大小
        self.resize(500,500)
        #设置窗口位置(下面配置的是居于屏幕中间)
        qr = self.frameGeometry()
        cp = QtWidgets.QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())
        #设置窗口标题和图标
        self.setWindowTitle('窗口标题')
        self.setWindowIcon(QtGui.QIcon('2.png'))
        #设置窗口提示
        self.setToolTip('窗口提示')
        #设置label信息
        self.label = QtWidgets.QLabel(self)
        self.label.setGeometry(QtCore.QRect(100, 10, 100, 60))
        self.label.setText('这是lable信息')
        self.label.setObjectName('label')
        # 设置label提示
        self.label.setToolTip('label提示')
        #设置输入框
        self.textbox = Qt.QLineEdit(self)
        self.textbox.resize(100, 20)
        self.textbox.move(100, 50)
        # 设置输入框提示
        self.textbox.setToolTip('输入框提示')
        #设置按钮
        self.btn =QtWidgets.QPushButton('按钮',self)
        self.btn.resize(100,20)
        self.btn.move(200,50)
        # 设置按钮样式
        self.btn.setStyleSheet("background-color: rgb(164, 185, 255);"
                          "border-color: rgb(170, 150, 163);"
                          "font: 75 12pt \"Arial Narrow\";"
                          "color: rgb(126, 255, 46);")
        # 设置按钮提示
        self.btn.setToolTip('按钮提示')
        #点击鼠标触发事件
        self.btn.clicked.connect(self.clickbtn)
        #展示窗口
        self.show()
    #点击鼠标触发函数
    def clickbtn(self):
        #打印出输入框的信息
        textboxValue = self.textbox.text()
        self.value = textboxValue
        panduan()
        QtWidgets.QMessageBox.question(self, "信息", '你输入的输入框内容为:' + textboxValue,QtWidgets.QMessageBox.Ok, QtWidgets.QMessageBox.Ok)
        #清空输入框信息
        self.textbox.setText('')
    #关闭窗口事件重写
    def closeEvent(self, QCloseEvent):
        reply = QtWidgets.QMessageBox.question(self, '警告',"确定关闭当前窗口?", QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No, QtWidgets.QMessageBox.No)
        if reply == QtWidgets.QMessageBox.Yes:
            QCloseEvent.accept()
        else:
            QCloseEvent.ignore()

def panduan():
    time.sleep(3)
    if get_data() == 1:
        print("[]")
    else:
        now_times = get_data()
        now_data = now_times.strftime("%Y.%m.%d")
        zzz()

if __name__ == '__main__':
    list_LHQ = get_LHQ()
    print('SAGE投料V1.0（状态：2）')
    # key = int(input('输入key:'))
    time.sleep(3)



    app = QtWidgets.QApplication(sys.argv)
    gui = GUI()
    sys.exit(app.exec_())



# if __name__ == "__main__":
#     list_LHQ = get_LHQ()
#     print('SAGE投料V1.0（状态：2）')
#     key = int(input('输入key:'))
#     time.sleep(3)
#     # sage_login()
#     while 2:
#         time.sleep(3)
#         if get_data() == 1:
#             print("[]")
#         else:
#             now_times = get_data()
#             now_data = now_times.strftime("%Y.%m.%d")
#             zzz()
